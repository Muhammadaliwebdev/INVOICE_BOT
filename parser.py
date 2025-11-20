# parser.py (Render-ready)
import os
import io
from pathlib import Path
from datetime import datetime
from typing import Optional, Union, List
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------
# Umumiy sozlamalar
# ---------------------------
# REPORTS_DIR ni environment variable orqali boshqaramiz (Render uchun qulay):
REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "reports"))
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

READ_KW = dict(header=None, engine="openpyxl")


# ---------------------------
# Kichik yordamchilar
# ---------------------------
def _to_str(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    return "" if s.lower() in ("nan", "none") else s


def _safe_iat(df: pd.DataFrame, r: int, c: int) -> str:
    try:
        return _to_str(df.iat[r, c])
    except Exception:
        return ""


def _coerce_date(s: str) -> str:
    if not s:
        return ""
    d = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(d):
        try:
            if isinstance(s, (datetime, pd.Timestamp)):
                d = pd.to_datetime(s)
        except Exception:
            pass
    if pd.isna(d):
        return s
    return d.strftime("%Y-%m-%d")


def _scan_after(df: pd.DataFrame, start_row: int, col: int, max_gap: int = 2) -> List[str]:
    out: List[str] = []
    gap = 0
    r = start_row
    nrows = len(df)
    while r < nrows:
        val = _to_str(df.iat[r, col]) if col < len(df.columns) else ""
        if val:
            out.append(val)
            gap = 0
        else:
            gap += 1
            if gap >= max_gap:
                break
        r += 1
    return out


def _open_ws(file_obj_or_path: Union[str, Path, io.BytesIO, io.BufferedReader]) -> Worksheet:
    wb = load_workbook(file_obj_or_path, data_only=True)
    ws = wb["Инвойс"] if "Инвойс" in wb.sheetnames else wb.active
    return ws


def _parse_number(val) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if not s:
        return None

    # faqat raqam, nuqta, vergul, minus
    s = re.sub(r"[^\d,.\-]", "", s)

    # ikkalasi bo‘lsa — oxirgi uchragan belgi decimal
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _find_header_and_cost_cols(ws: Worksheet) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Jadval sarlavhasini va 'Стоимость ...' ustunlarini topish.
    Returns: (header_row, col_cost, col_cost_vat)
    """
    max_r, max_c = ws.max_row, ws.max_column
    header_row = None
    col_cost = None
    col_cost_vat = None

    for r in range(1, min(max_r, 120) + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            row_vals.append("" if v is None else str(v).strip())
        row_join = " ".join(row_vals).lower()
        if any(k in row_join for k in ["наименование", "стоимость", "цена за", "ед. изм"]):
            header_row = r
            for c in range(1, max_c + 1):
                v = row_vals[c - 1].lower()
                if ("стоимость" in v and "ндс" in v) or "с учетом ндс" in v:
                    col_cost_vat = c
                elif "стоимость" in v and col_cost is None:
                    col_cost = c
            break

    return header_row, col_cost, col_cost_vat


def _find_itogo_row(ws: Worksheet) -> Optional[int]:
    """
    ИТОГО/ВСЕГО/Total qatordagi satrni pastdan yuqoriga qidiradi.
    """
    max_r, max_c = ws.max_row, ws.max_column
    keys = ("итого", "итого:", "итог", "всего", "total")
    for r in range(max_r, 0, -1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(k in v.lower() for k in keys):
                return r
    return None


def _find_invoice_total(ws: Worksheet) -> Optional[float]:
    """
    1) 'ИТОГО' qatorini topib, mos ustundan qiymatni oladi (VAT ustuni ustun).
    2) Agar topilmasa, 'Стоимость …' ustunini summalaydi (headerdan past, bo‘sh qator gap>=2 bo‘lguncha).
    """
    header_row, col_cost, col_cost_vat = _find_header_and_cost_cols(ws)
    itogo_row = _find_itogo_row(ws)

    max_c = ws.max_column

    def pick_from_row(row: int) -> Optional[float]:
        # ustuvorlik: VAT -> cost -> o‘sha qatordagi eng o‘ng numeric
        if col_cost_vat:
            v = _parse_number(ws.cell(row, col_cost_vat).value)
            if v is not None:
                return v
        if col_cost:
            v = _parse_number(ws.cell(row, col_cost).value)
            if v is not None:
                return v
        # fallback: o‘ngdan chapga qarab birinchi raqam
        for c in range(max_c, 0, -1):
            v = _parse_number(ws.cell(row, c).value)
            if v is not None:
                return v
        return None

    # 1) ИТОГО qatori bo‘lsa — bevosita o‘sha yerdan olamiz
    if itogo_row:
        v = pick_from_row(itogo_row)
        if v is not None:
            return v

    # 2) Summalash fallback
    if header_row and (col_cost_vat or col_cost):
        col = col_cost_vat or col_cost
        total = 0.0
        gap = 0
        has = False
        r = header_row + 1
        max_r = ws.max_row
        while r <= max_r:
            v = _parse_number(ws.cell(r, col).value)
            if v is None:
                gap += 1
                if gap >= 2:
                    break
            else:
                total += v
                has = True
                gap = 0
            r += 1
        if has:
            return total

    return None


# ---------------------------
# Asosiy parser
# ---------------------------
def extract_invoice_data(file_obj_or_path: Union[str, Path, io.BytesIO, io.BufferedReader]) -> dict:
    # Pandas DF (matnli qidiruvlar uchun)
    try:
        df = pd.read_excel(file_obj_or_path, sheet_name="Инвойс", **READ_KW)
    except Exception:
        df = pd.read_excel(file_obj_or_path, **READ_KW)
    df = df.fillna("")

    # Invoice raqami C1
    raw_inv = _safe_iat(df, 0, 2)
    invoice_number = (
        raw_inv.replace("ИНВОЙС №", "")
        .replace("ИНОЙС №", "")
        .replace("INVOICE №", "")
        .replace("INVOICE", "")
        .replace("от", "")
        .replace(":", "")
        .strip()
    )

    # Sana F1
    invoice_date = _coerce_date(_safe_iat(df, 0, 5))

    # Transport raqami E26
    vehicle_number = _safe_iat(df, 25, 4)

    # Firma nomi A5 (yoki topib olamiz)
    firm_name = _safe_iat(df, 4, 0)
    if not firm_name:
        found = ""
        for r in range(min(20, len(df))):
            for c in range(min(5, len(df.columns))):
                txt = _to_str(df.iat[r, c])
                if txt and any(k in txt for k in ["Фирма", "Поставщик", "Компания"]):
                    cand = _safe_iat(df, r + 1, c) or _safe_iat(df, r, c + 1)
                    if cand:
                        found = cand
                        break
            if found:
                break
        firm_name = found

    # Грузополучатель
    consignee: Optional[str] = None
    for r in range(len(df)):
        for c in range(len(df.columns)):
            cell = _to_str(df.iat[r, c])
            if cell.startswith("Грузополучатель") or cell.startswith("ГРУЗОПОЛУЧАТЕЛЬ"):
                cand = _safe_iat(df, r + 1, c) or _safe_iat(df, r, c + 1)
                consignee = cand if cand else ""
                break
        if consignee is not None:
            break

    # Tovarlar ro‘yxati (B30 dan)
    product_names = _scan_after(df, start_row=29, col=1, max_gap=2)

    # SUMMA: dinamik aniqlash
    ws = _open_ws(file_obj_or_path)
    invoice_sum = _find_invoice_total(ws)

    return {
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "vehicle_number": vehicle_number,
        "firm_name": firm_name,
        "consignee": consignee or "",
        "products": ", ".join(p for p in product_names if p),
        "invoice_sum": invoice_sum if invoice_sum is not None else "",
    }


# ---------------------------
# Hisobotni uslublash
# ---------------------------
def style_excel_report(filename: Union[str, Path]) -> None:
    wb = load_workbook(filename)
    ws = wb.active

    if ws.max_row >= 1:
        header_fill = PatternFill(start_color="0E3242", end_color="0E3242", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        1: 15,  # Sana
        2: 20,  # Invoice raqami
        3: 20,  # Transport raqami
        4: 25,  # Firma nomi
        5: 25,  # Qabul qiluvchi
        6: 25,  # Yuk tushirish joyi
        7: 20,  # Mijoz
        8: 40,  # Tovar nomi
        9: 18,  # Invoice summa
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # raqam formati
    try:
        sum_col_letter = get_column_letter(9)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{sum_col_letter}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    except Exception:
        pass

    ws.freeze_panes = "A2"
    wb.save(filename)


# ---------------------------
# Hisobotga yozish
# ---------------------------
def write_to_report(
    data: dict,
    delivery_place: str,
    customer_name: str,
) -> str:
    ym = datetime.now().strftime("%Y_%m")
    report_path = REPORTS_DIR / f"otschot_{ym}.xlsx"

    columns = [
        "Sana",
        "Invoice raqami",
        "Transport raqami",
        "Firma nomi",
        "Qabul qiluvchi",
        "Yuk tushirish joyi",
        "Mijoz",
        "Tovar nomi",
        "Invoice summa",
    ]

    if report_path.exists():
        try:
            df_existing = pd.read_excel(report_path)
        except Exception:
            df_existing = pd.DataFrame(columns=columns)
    else:
        df_existing = pd.DataFrame(columns=columns)

    for col in columns:
        if col not in df_existing.columns:
            df_existing[col] = ""
    df_existing = df_existing[columns]

    new_row = {
        "Sana": data.get("invoice_date", ""),
        "Invoice raqami": data.get("invoice_number", ""),
        "Transport raqami": data.get("vehicle_number", ""),
        "Firma nomi": data.get("firm_name", ""),
        "Qabul qiluvchi": data.get("consignee", ""),
        "Yuk tushirish joyi": (delivery_place or "").strip(),
        "Mijoz": (customer_name or "").strip(),
        "Tovar nomi": data.get("products", ""),
        "Invoice summa": data.get("invoice_sum", ""),
    }

    df_out = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
    df_out.to_excel(report_path, index=False)

    style_excel_report(report_path)
    return str(report_path)


# ---------------------------
# End of parser.py
# ---------------------------


# ---------------------------
# requirements.txt
# ---------------------------
# Paste these lines into a file named `requirements.txt` in the repo root.
# Versions may be adjusted if needed.

# aiogram v3 (use the specific stable version you have). Example below uses a commonly used beta version:
aiogram==3.0.0b7
pandas
openpyxl
python-dateutil

# If you use any other libs in bot.py, add them here (example: aiohttp, python-dotenv, etc.)
    